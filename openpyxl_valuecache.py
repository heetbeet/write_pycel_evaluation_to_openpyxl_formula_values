import openpyxl
from openpyxl import LXML
import openpyxl.cell._writer
from openpyxl.compat import safe_string
from openpyxl.xml.functions import whitespace
from types import SimpleNamespace
from contextlib import contextmanager
from openpyxl.workbook import Workbook
from typing import Dict, Tuple, List, Any, Iterator
from openpyxl import load_workbook
from monkeypatching import monkeypatch_module_object


_writer_cell_original = openpyxl.cell._writer.write_cell


def _write_cell_cache(
    cached_values: Dict[Tuple[str, str], Any],
    xf: Any,
    worksheet: Any,
    cell: Any,
    *args,
    **kwargs,
) -> None:
    """
    Custom openpyxl.worksheet._writer.write_cell wrapper function with cached
    values support.

    Args:
        cached_values: An additional argument. This is a dictionary containing
            cached cell values with a mapping of {(worksheet_name, cell_coordinate): value ...}
        xf: Original XML writer object passed by openpyxl.
        worksheet: The original worksheet object passed by openpyxl.
        cell: The original cell to write as passed by openpyxl.
        *args: Any additional args passed by openpyxl.
        **kwargs: Any additional kwargs passed by openpyxl.

    Returns:
        None
    """
    key = (worksheet.title, cell.coordinate)
    if key not in cached_values:
        return _writer_cell_original(xf, worksheet, cell, *args, **kwargs)

    value = cached_values[key]
    if LXML:
        raise NotImplementedError("LXML not supported with cached values")

        ## The following code will result in duplicate "v" elements
        # write_cell(xf, worksheet, cell, *args, **kwargs)
        # with xf.element("v"):
        #    xf.write(safe_string(value))

    else:
        from xml.etree.ElementTree import SubElement

        ns = SimpleNamespace()

        def write(el):
            ns.el = el

        _writer_cell_original(
            SimpleNamespace(write=write), worksheet, cell, *args, **kwargs
        )

        if (sub_el := ns.el.find("v")) is None:
            sub_el = SubElement(ns.el, "v")

        if isinstance(value, str):
            sub_el.text = value
            whitespace(sub_el)
            ns.el.set("t", "str")
        else:
            sub_el.text = safe_string(value)

        xf.write(ns.el)


@contextmanager
def _monkey_patch_openpyxl_write_cell(
    cached_values: Dict[Tuple[str, str], Any]
) -> Iterator[None]:
    """
    Context manager to monkeypatch the `write_cell` function in the openpyxl package.

    Args:
        cached_values (Dict[Tuple[str, str], Any]): A dictionary containing cached cell values.

    Yields:
        None: Yields control back to the caller, reverts changes upon exit.
    """

    def _write_cell_cached_closure(*args, **kwargs):
        return _write_cell_cache(cached_values, *args, **kwargs)

    with monkeypatch_module_object(
        openpyxl, _writer_cell_original, _write_cell_cached_closure
    ):
        yield


def save_workbook_with_cache(
    workbook: Workbook, filename: str, cached_values: Dict[Tuple[str, str], int]
) -> None:
    """
    Save an Openpyxl Workbook with certain cell values overridden by cached values.
    This allows you to prepopulate certain cells with values without having to
    calculate the workbook in an external application.

    This function saves a workbook after applying cached values to specific cells.
    The cached values are specified in a dictionary where the keys are tuples
    consisting of the worksheet name and cell coordinate, and the values are the
    cached values to apply.

    Args:
        workbook (Workbook): The Openpyxl Workbook object to save.
        filename (str): The name of the file to save the workbook as.
        cached_values (Dict[Tuple[str, str], int]): Dictionary containing the cached values.
            The keys are tuples where the first element is the worksheet name and the
            second element is the cell coordinate (e.g., ('Sheet1', 'A1')). The values
            are the cached values to set for those cells.

    Returns:
        None
    """

    with _monkey_patch_openpyxl_write_cell(cached_values):
        return workbook.save(filename)


if __name__ == "__main__":
    # Move this to tests or something

    cached_values = {}
    cached_values[("Sheet1", "A1")] = 9999

    workbook = load_workbook("your_workbook.xlsx")
    workbook.calculation.calcMode = "manual"
    save_workbook_with_cache(workbook, "your_workbook_cached.xlsx", cached_values)

    # openpyxl read cached workbook data only and see what A1 is
    workbook = load_workbook(
        "your_workbook_cached.xlsx", read_only=True, data_only=True
    )
    print(workbook["Sheet1"]["A1"].value)
