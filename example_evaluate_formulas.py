from pycel_valuecals import extract_formula_calculations
from openpyxl_replacetable import replace_table
from openpyxl_valuecache import save_workbook_with_cache
import re

import pandas as pd
import openpyxl

# Load the Excel workbook
wb = openpyxl.load_workbook("va-template.xlsx")

def fill_in_submission(wb, value_mapping):
    sheet = wb["Submission"]

    # Clear
    for col in sheet.iter_cols(min_col=3):
        for cell in col:
            cell.value = None

    ref_map = {
        str(i.value).lower().replace(" ", "_"): int(i.coordinate.strip("ABCDEFGHIJKLMNOPQRSTUVWXYZ")) for i in sheet["B"] if i.value
    }

    for key, value in value_mapping.items():
        if isinstance(value, (tuple, list)):
            for i, v in enumerate(value):
                sheet.cell(row=ref_map[key], column=3 + i).value = v
        else:
            sheet.cell(row=ref_map[key], column=3).value = value
        


def calculate_and_prune_summary(wb):
    sheet2 = wb["Summary"]

    cell_values = extract_formula_calculations(wb)

    empty_cols = []
    for i, col in enumerate(sheet2.iter_cols()):
        cell_is_filled = False
        for cell in col:
            if cell_values.get((sheet2.title, cell.coordinate), cell.value) not in (
                None,
                "",
            ):
                cell_is_filled = True
                break

        if not cell_is_filled and i > 0:
            empty_cols.append(i + 1)

    for i in empty_cols[::-1]:
        sheet2.delete_cols(i)

    # go to the second sheet, and iterate through all the rows:
    # cell_values = extract_formula_calculations(wb)
    empty_rows = []
    for i, row in enumerate(sheet2.iter_rows()):
        # go through each cell in the row
        cell_is_filled = False
        for cell in row:
            if cell_values.get((sheet2.title, cell.coordinate), cell.value) not in (
                None,
                "",
            ):
                cell_is_filled = True
                break
        if not cell_is_filled and i > 0:
            empty_rows.append(i + 1)

    for i in empty_rows[::-1]:
        sheet2.delete_rows(i)

    # find the last row and column with anything (formula or value) in it
    max_row = sheet2.max_row

    regex = re.compile(r"([A-Z]+)(\d+):([A-Z]+)(\d+)")

    def replace_m(match):
        new_end_row = str(min(max_row - 1, int(match.group(4))))
        return f"{match.group(1)}{match.group(2)}:{match.group(3)}{new_end_row}"

    for cell in sheet2[max_row]:
        # fix all formula slices in the last row not to include the deleted rows
        if str(cell.value).startswith("="):
            cell.value = regex.sub(replace_m, cell.value)

    # Add a double border to the last row
    for i, cell in enumerate(sheet2[max_row - 1]):
        if i > 0 and i <= sheet2.max_column - 1:
            cell.border = openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(border_style="double")
            )

    return extract_formula_calculations(wb)


# Create DataFrame
data = {
    "Virtual Actuary name": [
        "A",
        "B",
        "A",
        "C",
        "D",
        "B",
        "A",
        "C",
        "D",
        "A",
    ],
    "Virtual Actuary ID": [
        "VA-001",
        "VA-002",
        "VA-001",
        "VA-003",
        "VA-004",
        "VA-002",
        "VA-001",
        "VA-003",
        "VA-004",
        "VA-001",
    ],
    "client": [
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
        "Sanlam",
    ],
    "project": [
        "SEM_GI_Support",
        "SEM_GI_TM1Testing",
        "SEM_GI_Support",
        "SEM_GI_ActuarialBAU",
        "SEM_GI_Support",
        "SEM_GI_TM1Testing",
        "SEM_GI_Support",
        "SEM_GI_TM1Testing",
        "SEM_GI_ActuarialBAU",
        "SEM_GI_ActuarialBAU",
    ],
    "tags": [
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "week_starting": [
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
        "2023-06-26",
    ],
    "description": [
        "May run",
        "May run",
        "May run",
        "May run",
        "May run",
        "May run",
        "May run",
        "May run",
        "May run",
        "May run",
    ],
    "date": [
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
        "2023-06-28",
    ],
    "duration": [
        1.32,
        2.51,
        0.80,
        2.89,
        1.5,
        0.7,
        1.2,
        1.0,
        1.3,
        0.9,
    ],
    "retrieval_date": [
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
        "2023-06-30",
    ],
}


df = pd.DataFrame(data)

# Fill in submission
fill_in_submission(wb, {
    "template_version": "V1.0.0",
    "invoice_key": "INV-COM-00001",
    "va_entity": "Virtual Actuary (Pty) Ltd",
    "client_master_name": "Company",
    "va_champion": "A",
    "start_date": "2023-06-09",
    "end_date": "2023-06-28",
    "project_names": ["SEM_GI_Support", "SEM_GI_TM1Testing", "SEM_GI_ActuarialBAU"],
    "virtual_actuary_names": ["A", "B", "C", "D"],
    "virtual_actuary_id": ["VA-001", "VA-002", "VA-003", "VA-004"],
})

# Fill in table
replace_table(wb, "Timesheet", df)

# Calculate and prune summary
cell_values = calculate_and_prune_summary(wb)

wb.calculation.calcMode = "manual"
save_workbook_with_cache(wb, "va-template-cached.xlsx", cell_values)
