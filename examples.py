from locate import this_dir
from openpyxl import load_workbook
from pycel import ExcelCompiler
from pathlib import Path

def evaluate_and_test(filename):
    filename = Path(filename)
    # Step 1: Load the Excel workbook using openpyxl
    wb = load_workbook(filename=filename)

    # Step 2: Change cell A1 to 10 in the first sheet
    first_sheet_name = wb.sheetnames[0]
    ws = wb[first_sheet_name]
    ws['A1'].value = 10
    
    # Step 3: Compile the workbook using pycel
    compiler = ExcelCompiler(excel=wb)

    # compiler to evaluate A2
    print(compiler.evaluate("A2"))

    # Step 4: Read the value of cell A2 in the first sheet
    new_value_b1 = ws['A2'].value
    print(f"New value of A2 in the first sheet: {new_value_b1}")

    # Optionally, save the updated and evaluated workbook back to an .xlsx file
    
    wb.save(filename.parent / ("evaluated_" + filename.name))

# Usage example
evaluate_and_test(this_dir() / "tester.xlsx")
evaluate_and_test(this_dir() / "tester1.xlsx")


# open tester1 with openpyxl and read the value of A2
print(f"Value of A2 in the first sheet: {(wb:=load_workbook(filename=this_dir() / 'tester1.xlsx', data_only=True))[wb.sheetnames[0]]['A2'].value}")

