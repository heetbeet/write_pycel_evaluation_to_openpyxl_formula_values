import openpyxl
from openpyxl import LXML
from openpyxl.cell import _writer as _cell_writer
from openpyxl.compat import safe_string
from types import SimpleNamespace
from contextlib import contextmanager
from openpyxl.workbook import Workbook
from typing import Dict, Tuple, List, Any, Iterator, Union
from openpyxl import load_workbook
from pathlib import Path


def _is_subpath(child_path: Union[Path, str], parent_path: Union[Path, str]) -> bool:
    """
    Check if one path is a subpath of another.

    Args:
        child_path: The child path to check.
        parent_path: The parent path to check against.

    Returns:
        True if child_path is a subpath of parent_path, False otherwise.
    """
    child = Path(child_path).resolve()
    parent = Path(parent_path).resolve()
    try:
        child.relative_to(parent)
        return True
    except ValueError:
        return False


def _module_file_path(module: Any) -> Union[None, Path]:
    """
    Retrieve the file path of a module.

    Args:
        module: The module whose file path is to be retrieved.

    Returns:
        The file path of the module if it exists, otherwise None.
    """
    try:
        return Path(module.__file__).resolve()
    except AttributeError:
        return None


def _find_submodules(module: Any) -> Dict[Path, Any]:
    """
    Recursively find all submodules of a given module.

    Args:
        module: The parent module to search from.

    Returns:
        A dictionary mapping file paths to module objects.

    Raises:
        ValueError: If the module is not loaded from a file.
    """
    cache_dict = {}

    module_path = _module_file_path(module)
    if module_path is None:
        raise ValueError("Module must be a module loaded from a file.")

    module_dir = module_path.parent

    def recurse(module):
        path = _module_file_path(module)
        if (path is None) or (path in cache_dict) or not _is_subpath(path, module_dir):
            return

        cache_dict[path] = module

        for attr_name in dir(module):
            try:
                attr = getattr(module, attr_name)
            except Exception:
                continue

            recurse(attr)

    recurse(module)
    return cache_dict


def _list_monkeypatch_locations(package: Any, function: Any) -> List[Tuple[Any, str]]:
    """
    List all locations where a given function is used in a package.

    Args:
        package: The package in which to search for the function.
        function: The function to look for.

    Returns:
        A list of tuples containing module and attribute name.
    """
    locations = _find_submodules(package)

    f_locations = []
    for filename, module in locations.items():
        for attr_name in dir(module):
            if (attr := getattr(module, attr_name)) is function:
                f_locations.append((module, attr_name))

    return f_locations


_write_cell_orig = _cell_writer.write_cell


def _write_cell_cache(
    cached_values: Dict[Tuple[str, str], Any],
    xf: Any,
    worksheet: Any,
    cell: Any,
    *args,
    **kwargs,
) -> None:
    """
    Custom openpyxl.worksheet._writer.write_cell wrapper function with cached
    values support.

    Args:
        cached_values: An additional argument. This is a dictionary containing
            cached cell values with a mapping of {(worksheet_name, cell_coordinate): value ...}
        xf: Original XML writer object passed by openpyxl.
        worksheet: The original worksheet object passed by openpyxl.
        cell: The original cell to write as passed by openpyxl.
        *args: Any additional args passed by openpyxl.
        **kwargs: Any additional kwargs passed by openpyxl.

    Returns:
        None
    """
    key = (worksheet.title, cell.coordinate)
    if key not in cached_values:
        return _write_cell_orig(xf, worksheet, cell, *args, **kwargs)

    value = cached_values[key]
    if LXML:
        raise NotImplementedError("LXML not supported with cached values")

        ## The following code will result in duplicate "v" elements
        # _write_cell_orig(xf, worksheet, cell, *args, **kwargs)
        # with xf.element("v"):
        #    xf.write(safe_string(value))

    else:
        from xml.etree.ElementTree import SubElement

        ns = SimpleNamespace()

        def write(el):
            ns.el = el

        _write_cell_orig(SimpleNamespace(write=write), worksheet, cell, *args, **kwargs)

        # Add cached value to the "v" element in the cell's XML
        v_elem = ns.el.find("v")
        if v_elem is None:
            SubElement(ns.el, "v")

        v_elem.text = safe_string(value)
        xf.write(ns.el)


@contextmanager
def _monkey_patch_openpyxl_write_cell(
    cached_values: Dict[Tuple[str, str], Any]
) -> Iterator[None]:
    """
    Context manager to monkeypatch the `write_cell` function in the openpyxl package.

    Args:
        cached_values (Dict[Tuple[str, str], Any]): A dictionary containing cached cell values.

    Yields:
        None: Yields control back to the caller, reverts changes upon exit.
    """
    try:
        # Unfortunately we need to monkeypatch openpyxl everywhere where a `from ... import write_cell` is done
        def _write_cell_closure(*args, **kwargs):
            return _write_cell_cache(cached_values, *args, **kwargs)

        f_list = _list_monkeypatch_locations(openpyxl, _write_cell_orig)

        for module, function_name in f_list:
            setattr(module, function_name, _write_cell_closure)

        yield
    finally:
        for module, function_name in f_list:
            setattr(module, function_name, _write_cell_orig)


def save_workbook_with_cache(
    workbook: Workbook, filename: str, cached_values: Dict[Tuple[str, str], int]
) -> None:
    """
    Save an Openpyxl Workbook with certain cell values overridden by cached values.
    This allows you to prepopulate certain cells with values without having to
    calculate the workbook in an external application.

    This function saves a workbook after applying cached values to specific cells.
    The cached values are specified in a dictionary where the keys are tuples
    consisting of the worksheet name and cell coordinate, and the values are the
    cached values to apply.

    Args:
        workbook (Workbook): The Openpyxl Workbook object to save.
        filename (str): The name of the file to save the workbook as.
        cached_values (Dict[Tuple[str, str], int]): Dictionary containing the cached values.
            The keys are tuples where the first element is the worksheet name and the
            second element is the cell coordinate (e.g., ('Sheet1', 'A1')). The values
            are the cached values to set for those cells.

    Returns:
        None
    """

    with _monkey_patch_openpyxl_write_cell(cached_values):
        return workbook.save(filename)


if __name__ == "__main__":
    # Move this to tests or something

    cached_values = {}
    cached_values[("Sheet1", "A1")] = 9999

    workbook = load_workbook("your_workbook.xlsx")
    workbook.calculation.calcMode = "manual"
    save_workbook_with_cache(workbook, "your_workbook_cached.xlsx", cached_values)

    # openpyxl read cached workbook data only and see what A1 is
    workbook = load_workbook(
        "your_workbook_cached.xlsx", read_only=True, data_only=True
    )
    print(workbook["Sheet1"]["A1"].value)
