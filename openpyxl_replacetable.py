from openpyxl.utils import coordinate_to_tuple
from openpyxl.workbook import Workbook
import pandas as pd

class TableNotFoundError(Exception):
    pass


def replace_table(wb: Workbook, tablename: str, df: pd.DataFrame) -> None:
    """
    Replaces the data in an existing Excel table with new data from a Pandas DataFrame.
    If the table does not exist, raises a TableNotFoundException.
    """

    table = None
    ws = None

    # Loop through all worksheets to find the table
    for sheet in wb:
        for tbl in sheet._tables.values():
            if tbl.name == tablename:
                table = tbl
                ws = sheet
                break
        if table:
            break

    # If table is not found, throw an exception
    if table is None:
        raise TableNotFoundError(f'Table not found: "{tablename}"')

    # Parse table.ref string to get the start and end cells
    start_cell, end_cell = table.ref.split(":")
    start_row, start_col = coordinate_to_tuple(start_cell)
    end_row, end_col = coordinate_to_tuple(end_cell)

    # Clear existing data
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.value = None

    # Update table headers and data
    new_headers = list(df.columns)
    new_data = df.values.tolist()
    all_data = [new_headers] + new_data

    # Write all data to the worksheet
    for i, row_data in enumerate(all_data, start=start_row):
        for j, value in enumerate(row_data, start=start_col):
            ws.cell(row=i, column=j, value=value)

    # Resize table to fit new data
    end_row = start_row + len(all_data) - 1
    end_col = start_col + len(new_headers) - 1
    new_range = (
        ws.cell(row=start_row, column=start_col).coordinate
        + ":"
        + ws.cell(row=end_row, column=end_col).coordinate
    )
    table.ref = new_range
